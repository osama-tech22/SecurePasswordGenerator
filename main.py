from tkinter import *
import pyperclip
from openpyxl import Workbook, load_workbook
from datetime import datetime
import secrets
import string

root = Tk()
root.title('Simo Password Generator')
root.geometry("450x450")


# Secure password generation using secrets
def generate_secure_password(length=12):
    characters = string.ascii_letters + string.digits + string.punctuation
    password = ''.join(secrets.choice(characters) for i in range(length))
    return password


def save_to_excel(website, password):
    filename = "passwords.xlsx"
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Website", "Date and Time", "Password"])

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([website, current_time, password])
    wb.save(filename)


def generate_password():
    pw_entry.delete(0, END)
    pw_length = int(my_entery2.get())
    my_password = generate_secure_password(pw_length)
    pw_entry.insert(0, my_password)

    website_name = my_entry1.get()
    if website_name:
        save_to_excel(website_name, my_password)


def clipper():
    root.clipboard_clear()
    root.clipboard_append(pw_entry.get())


lf1 = LabelFrame(root, text="Website")
lf1.pack(pady=20)
lf2 = LabelFrame(root, text="Password Length")
lf2.pack(pady=20)

my_entry1 = Entry(lf1, font=("Helvetica", 24))
my_entry1.pack(pady=20, padx=20)
my_entery2 = Entry(lf2, font=("Helvetica", 24))
my_entery2.pack(pady=20, padx=20)

pw_entry = Entry(root, text="", font=("Helvetica", 24), bd=0, bg="gray")
pw_entry.pack(pady=20)

my_frame = Frame(root)
my_frame.pack(pady=20)

my_button = Button(my_frame, text="Generate Password", command=generate_password)
my_button.grid(row=0, column=0, padx=10)

clip_button = Button(my_frame, text="Copy To Clipboard", command=clipper)
clip_button.grid(row=0, column=1, padx=10)

root.mainloop()
